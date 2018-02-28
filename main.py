import sys
from openpyxl import load_workbook
import requests
from pprint import pprint

KEY = "Default Google Maps API Key"
distances=[]

def getLeg(zip1, zip2):
    url = "https://maps.googleapis.com/maps/api/directions/json?origin=" + zip1 + "&destination=" + zip2 + "&sensor=false" + "&key=" + KEY
    resp = requests.get(url)
    resp_json = resp.json()

    try:
        route = resp_json['routes'][0]
        leg = route['legs'][0]
        return leg
    except Exception as e:
        print(pprint(resp_json))
        return "Error"


if __name__ == "__main__":
    print("This program when given an excel sheet, looks at the first 2 coloumns of the first sheet.\nAssuming they are" \
          " both zip codes, it finds the distance between them using the google maps api\nand puts the distance in the 3rd" \
          " column. It then saves the file.\nIf the defualt api key does not allow enough api requests\nyou can set the " \
          "api key by adding it as a second parameter\n\n" \
          "Example usage: `python main.py dragged/and/dropped/filename apikey`\n")
    if len(sys.argv) > 1:
        filename = sys.argv[1]
    else:
        filename = input("Drag and drop the file name: ").rstrip()

    if len(sys.argv) > 2:
        KEY = sys.argv[2]

    try:
        wb = load_workbook(filename=filename)
    except Exception:
        print("Could not open file")
        exit()

    ws = wb.active
    done = False
    i = 0

    while not done:
        i = i + 1
        zipa = str(ws["A" + str(i)].value)
        zipb = str(ws["B" + str(i)].value)

        if zipa == 'None' or zipb == 'None':
            print("Finished at row " + str(i))
            done = True
        else:

            while len(zipa) < 5:
                zipa = "0" + zipa

            while len(zipb) < 5:
                zipb = "0" + zipb

            leg = getLeg(zipa, zipb)
            ws['C' + str(i)] = leg['distance']['text']
            ws['D' + str(i)] = leg['duration']['text']
            distances.append(leg['distance']['value'])
            print(zipa + " " + zipb + " : " + leg['distance']['text'])

    ws['E1'] = "Average distance in meters"
    ws['E2'] = sum(distances) / float(len(distances))

    ws['F1'] = "Average distance in miles"
    ws['F2'] = ws['E2'].value * 0.000621371

    wb.save(filename)
